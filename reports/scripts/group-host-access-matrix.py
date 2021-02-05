import pymongo
import json
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import sys

with open(sys.argv[1]+'etc/config.json') as config_file:
    config = json.load(config_file)

client = pymongo.MongoClient(config['mongo']['url'])
mdb = client["MAICS"]
access_groups = mdb["access_groups"]
groups = mdb["groups"]
hosts = mdb["hosts"]
users = mdb["users"]

row, col = 1,1
### Sheet format fars
#fonts
bold = Font(bold=True)
#colors
fillBlue = PatternFill(start_color=Color('618DD3'),
               end_color=Color('3C82F2'),
               fill_type='solid')
fillGreen = PatternFill(start_color=Color('51B716'),
                end_color=Color('3C82F2'),
                fill_type='solid')
fillRed = PatternFill(start_color=Color('FF245B'),
                end_color=Color('3C82F2'),
                fill_type='solid')
fillYellow = PatternFill(start_color=Color('FFDD00'),
                end_color=Color('FFDD00'),
                fill_type='solid')
#Alignments
center = Alignment(horizontal='center')

#create xlsx
book = Workbook()
sheet = book.create_sheet("MAICS Access")

while col < 500:
    i = get_column_letter(col)
    sheet.column_dimensions[i].width = 20
    col += 1

col = 1
sheet.column_dimensions['A'].width
sheet.cell(row, col).value = "Sheet last update"
col += 1
sheet.cell(row, col).value = datetime.now()
col = 1
row += 2
sheet.cell(row, col).value = "Users w/ system-wide access"
sheet.cell(row, col).fill = fillBlue

for a in users.find( { "role": "admin" }, {"name":1, "surname":1}).sort("surname"):
    row +=1
    sheet.cell(row, col).value = a['surname']+" "+a['name']

col = 1
row += 2
sheet.cell(row, col).value = "Access matrix"
sheet.cell(row, col).font = bold
sheet.cell(row, col).fill = fillBlue
row += 1
sheet.cell(row, col).value = "group/host"
sheet.cell(row, col).font = bold

col = 2
for g in groups.find( {}, {"name":1}).sort("name"):
    sheet.cell(row, col).value=g['name']
    col+=1

for h in hosts.find({}).sort("hostname"):
    bitmap = []
    col = 1
    row +=1
    sheet.cell(row, col).value=h['hostname']
    col +=1
    for g in groups.find( {}, {"name":1}).sort("name"):
        acl = access_groups.count_documents({"group": g['name'], "hostname": h['hostname']})
        bitmap.append(acl)
    for b in bitmap:
        if(b == 1):
            sheet.cell(row, col).fill = fillGreen
        else:
            sheet.cell(row, col).fill = fillRed
        col += 1

book.remove(book['Sheet'])
book.save(config['maics']['dir']+"reports/group-host-access-matrix.xlsx")
